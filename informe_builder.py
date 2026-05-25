"""
Genera libros Excel en el formato del informe de lecturas (plantilla + ejemplo enviado)
a partir del archivo plano exportado por el administrador.
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, replace
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

INVALID_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')

from admin_columns import (
    GRADO_COL,
    GRUPO_COL,
    INST_COL,
    REQUIRED_ADMIN_COLS,
    normalize_admin_dataframe_columns,
)


@dataclass
class BuildConfig:
    """Parámetros del informe generado."""

    institution_contains: str = ""
    reading_title: str = ""
    report_date: str = ""
    # Encabezado de plantilla FORMATO_LEC (I2=asignatura, I3=docente, I4=elaborado por)
    asignatura: str = ""
    docente: str = ""
    elaborado_por: str = ""
    # Nombre de la institución en la plantilla (celda D2); útil con export por colegio.
    institution_label: str = ""
    otros_sheet_name: str = "OTROS"
    # Si True: solo filas de la institución indicada; OTROS = solo grado/grupo no clasificable.
    # Si False: el resto de instituciones va a OTROS (útil para un solo archivo maestro).
    only_primary_institution: bool = True


def _normalize_grade(grade) -> str | None:
    if grade is None or (isinstance(grade, float) and pd.isna(grade)):
        return None
    s = str(grade).strip()
    s = s.replace("º", "").replace("°", "").strip()
    # quitar texto tipo "Estudiantes PFC (ENSAS)"
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits:
        return digits
    # fallback: primera secuencia alfanumérica útil
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else None


def _normalize_subgroup(sub) -> str | None:
    if sub is None or (isinstance(sub, float) and pd.isna(sub)):
        return None
    s = str(sub).strip()
    # Valores tipo "-5" en el Excel suelen ser error de captura; no fusionar con el grupo "5".
    if re.match(r"^-\d+$", s):
        return None
    if not s:
        return None
    # "04" -> "4", pero mantener "A", "B"
    if re.fullmatch(r"\d+", s):
        return str(int(s))
    return s


def sheet_key_from_row(grade, subgroup) -> str | None:
    g = _normalize_grade(grade)
    sub = _normalize_subgroup(subgroup)
    if not g or not sub:
        return None
    return f"{g}-{sub}"


def safe_sheet_title(name: str, max_len: int = 31) -> str:
    name = INVALID_SHEET_CHARS.sub("-", name.strip())
    name = re.sub(r"-{2,}", "-", name).strip("-")
    if not name:
        name = "HOJA"
    return name[:max_len]


def _data_sheet_tab_basename(grade_group_key: str) -> str:
    """
    Títulos que empiezan por dígito (p. ej. «4-a») han dado problemas con openpyxl/Excel
    (colisiones por mayúsculas y búsquedas por nombre). Se prefija «G_» (grupo).
    """
    s = safe_sheet_title(grade_group_key, max_len=28)
    return f"G_{s}"[:31]


def _unique_sheet_title(base: str, used_lower: set[str], max_len: int = 31) -> str:
    """
    Excel no distingue mayúsculas en nombres de hoja; «4-A» y «4-a» colisionan y openpyxl renombra
    sin avisar (p. ej. a «4-a1»), lo que rompe el rellenado. Garantiza un nombre único.
    """
    title = safe_sheet_title(base, max_len=max_len)
    if title.lower() not in used_lower:
        used_lower.add(title.lower())
        return title
    n = 2
    while True:
        suffix = f" ({n})"
        stem = safe_sheet_title(base, max_len=max(1, max_len - len(suffix)))
        candidate = (stem + suffix)[:max_len]
        if candidate.lower() not in used_lower:
            used_lower.add(candidate.lower())
            return candidate
        n += 1


def _institution_match(cell_value, needle: str) -> bool:
    if needle is None or str(needle).strip() == "":
        return True
    if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
        return False
    return str(needle).lower() in str(cell_value).lower()


def _is_zip_xlsx(data: bytes) -> bool:
    return len(data) >= 2 and data[0:2] == b"PK"


def _sniff_data_kind(data: bytes, filename: str) -> str:
    """'csv' | 'xlsx' según extensión o cabecera del archivo."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if _is_zip_xlsx(data):
        return "xlsx"
    # Heurístico: texto con separador (exportaciones CSV)
    head = data[:16384] if len(data) > 16384 else data
    if b"\x00" in head:
        return "xlsx"  # last resort, fallará read_excel si no es
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            t = data.decode(enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
        first = t[:4000]
        if "\n" in first and ("," in first or ";" in first or "\t" in first):
            return "csv"
    return "xlsx"


def _read_csv_to_dataframe(data: bytes) -> pd.DataFrame:
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = data.decode(enc)
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
            continue
        for kwargs in (
            {"sep": None, "engine": "python"},
            {"sep": ";"},
            {"sep": ","},
            {"sep": "\t"},
        ):
            try:
                sio = io.StringIO(text)
                return pd.read_csv(sio, header=0, **kwargs)
            except Exception as e:
                last_err = e
                continue
    if last_err:
        raise ValueError("No se pudo leer el CSV (codificación o separador).") from last_err
    raise ValueError("No se pudo leer el CSV.")


def load_admin_dataframe(
    source: str | Path | BinaryIO,
    source_filename: str = "",
) -> tuple[pd.DataFrame, list[str]]:
    """
    Lee Excel o CSV del admin; devuelve DataFrame y lista de nombres de columnas.
    `source_filename` se usa con BytesIO para detectar .csv (el buffer no trae nombre).
    """
    if isinstance(source, (str, Path)):
        path = Path(source)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            df = _read_csv_to_dataframe(path.read_bytes())
        else:
            df = pd.read_excel(path, header=0, engine="openpyxl")
    else:
        if isinstance(source, (bytes, bytearray)):
            data = bytes(source)
        else:
            data = source.read()
        kind = _sniff_data_kind(data, source_filename)
        if kind == "csv":
            df = _read_csv_to_dataframe(data)
        else:
            df = pd.read_excel(io.BytesIO(data), header=0, engine="openpyxl")

    cols = [str(c).strip() for c in df.columns.tolist()]
    df.columns = cols
    df = normalize_admin_dataframe_columns(df)
    _validate_admin_dataframe_for_report(df)
    return df, list(df.columns)


def _validate_admin_dataframe_for_report(df: pd.DataFrame) -> None:
    """
    Cada formulario/lectura puede añadir columnas de preguntas distintas; lo imprescindible
    es conservar en el encabezado lo que Gravity exporta para grado, grupo e institución.
    """
    missing = [c for c in REQUIRED_ADMIN_COLS if c not in df.columns]
    if not missing:
        return
    preview = list(df.columns)[:32]
    label = "columnas" if len(missing) > 1 else "la columna"
    names = " ".join(f"«{c}»" for c in missing)
    raise ValueError(
        f"Falta {label} {names} (nombres exactos del export de Gravity u otro pliego). "
        "Las distintas lecturas cambian el texto de las preguntas, pero el bloque básico "
        "(Fecha, nombre, correo, institución, grado, número/letra de grupo) debe mantenerse. "
        f"Primeras columnas leídas: {preview!r}"
    )


def _pick_template_master_sheet_name(sheetnames: list[str]) -> str:
    if not sheetnames:
        raise ValueError("El Excel de plantilla no tiene hojas.")
    if "10-1" in sheetnames:
        return "10-1"
    if len(sheetnames) == 1:
        return sheetnames[0]
    first_few = ", ".join(f"«{n}»" for n in sheetnames[:5])
    more = f" (y {len(sheetnames) - 5} más)" if len(sheetnames) > 5 else ""
    raise ValueError(
        f"El archivo de plantilla no contiene la hoja base «10-1» y hay {len(sheetnames)} hojas ({first_few}{more}). "
        "Incluye o renombra la hoja base a «10-1» (como en la plantilla de ejemplo) o deja en el archivo **solo** la hoja con el cuerpo del informe. "
        "No subas en «plantilla» el Excel de **respuestas** del formulario: ese documento se elige en «Excel administrador»; "
        "cada lección/lectura puede exportarse con título de hoja distinto (p. ej. un cuento), pero el rol del archivo (datos frente a maqueta) es distinto."
    )


def _worksheet_row1_looks_like_gravity_admin_header(ws) -> bool:
    """Distingue la tabla de respuestas (export administrador) de la maqueta del informe."""
    max_c = min(ws.max_column, 32)
    parts: list[str] = []
    for c in range(1, max_c + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            parts.append(str(v).lower().strip())
    if len(parts) < 3:
        return False
    joined = " ".join(parts)
    if "institución o colegio" not in joined or "grado" not in joined:
        return False
    if "escribe tu nombre" in joined or "fecha" in joined or "correo" in joined:
        return True
    return False


def split_rows_by_sheet(
    df: pd.DataFrame,
    config: BuildConfig,
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Separa filas en:
    - buckets por grado-grupo (ej. '7-2') para institución principal
    - DataFrame OTROS: según configuración, filas no clasificables y/o otras instituciones
    """
    inst_col = INST_COL
    grado_col = GRADO_COL
    grupo_col = GRUPO_COL

    for c in (inst_col, grado_col, grupo_col):
        if c not in df.columns:
            raise KeyError(f"Falta la columna obligatoria «{c}». Columnas: {list(df.columns)}")

    work = df
    if config.only_primary_institution:
        mask = df[inst_col].apply(lambda v: _institution_match(v, config.institution_contains))
        work = df[mask].copy()
        otros_mask = pd.Series(False, index=work.index)
        # Si el texto de institución no coincide con ninguna fila, el informe quedaba vacío
        # mientras la vista previa en el front muestra todo. Usar todas las filas como respaldo.
        if work.empty and not df.empty:
            work = df.copy()
            otros_mask = pd.Series(False, index=work.index)
    else:
        otros_mask = ~work[inst_col].apply(lambda v: _institution_match(v, config.institution_contains))

    buckets: dict[str, list] = {}
    otros_extra: list = []

    for idx in work.index:
        if otros_mask.loc[idx] if idx in otros_mask.index else False:
            continue
        row = work.loc[idx]
        key = sheet_key_from_row(row[grado_col], row[grupo_col])
        if key is None:
            otros_extra.append(idx)
            continue
        buckets.setdefault(key, []).append(idx)

    sheet_frames: dict[str, pd.DataFrame] = {}
    used_titles: set[str] = set()
    for key, idx_list in sorted(buckets.items()):
        t = _unique_sheet_title(_data_sheet_tab_basename(key), used_titles)
        sheet_frames[t] = work.loc[idx_list].copy()

    otros_parts: list[pd.DataFrame] = []
    if not config.only_primary_institution:
        otros_parts.append(work.loc[otros_mask])
    if otros_extra:
        otros_parts.append(work.loc[otros_extra])
    otros_df = (
        pd.concat(otros_parts, axis=0).drop_duplicates()
        if otros_parts
        else pd.DataFrame(columns=work.columns)
    )

    return sheet_frames, otros_df


def unique_institution_values(df: pd.DataFrame) -> list[str]:
    """Valores no vacíos y únicos de «Institución o colegio», ordenados para export estable."""
    if INST_COL not in df.columns:
        raise KeyError(
            f"Falta la columna obligatoria «{INST_COL}» para listar colegios. Columnas: {list(df.columns)}",
        )
    s = df[INST_COL].dropna().map(lambda x: str(x).strip())
    s = s[s != ""]
    if s.empty:
        return []
    u = list(dict.fromkeys(s.tolist()))
    u.sort(key=lambda x: (x.lower(), x))
    return u


def _set_cell_merged_aware(ws, cell_ref: str, value) -> None:
    """Escribe en una celda; si participa en un rango fusionado, pone el valor en la esquina superior izquierda."""
    row, col = coordinate_to_tuple(cell_ref)
    for cr in ws.merged_cells.ranges:
        if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
            top_left = f"{get_column_letter(cr.min_col)}{cr.min_row}"
            ws[top_left] = value
            return
    ws[cell_ref] = value


def _safe_file_stem(name: str, max_len: int = 72) -> str:
    s = re.sub(r'[<>:"/\\|?*\n\r\t]+', "-", str(name).strip())
    s = re.sub(r"\s+", " ", s).strip() or "colegio"
    return s[:max_len]


def _compose_excel_bytes(
    template_path: Path,
    config: BuildConfig,
    header_row: list[str],
    sheet_frames: dict[str, pd.DataFrame],
    otros_df: pd.DataFrame,
    reference_df: pd.DataFrame,
) -> bytes:
    if not template_path.is_file():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    wb = load_workbook(template_path)
    master_name = _pick_template_master_sheet_name(wb.sheetnames)
    for sn in list(wb.sheetnames):
        if sn != master_name:
            wb.remove(wb[sn])

    if _worksheet_row1_looks_like_gravity_admin_header(wb[master_name]):
        raise ValueError(
            "Ese archivo parece un export de respuestas (administrador o Gravity), no el Excel de plantilla del informe. "
            "En la primera fila se ven encabezados como «Fecha», «Escribe tu nombre» e «Institución o colegio» (tabla de respuestas; "
            "cambia el cuento o las preguntas, pero el formato de export sigue similar). "
            "Úsalo como «Excel administrador». La plantilla (opcional) es el .xlsx con la maqueta: celdas B7, D2, I2 a I4, datos desde fila 8, "
            "hoja «10-1» o un archivo con una sola hoja con esa disposición."
        )

    ordered_sheets: list[str] = []
    for k in sorted(sheet_frames.keys(), key=lambda x: _sheet_sort_key(x)):
        ordered_sheets.append(k)
    otros_title = f"G_{safe_sheet_title(config.otros_sheet_name)}"[:31]
    if not otros_df.empty:
        ordered_sheets.append(otros_title)

    if not ordered_sheets:
        ordered_sheets = [otros_title]
        sheet_frames = {}
        if otros_df.empty:
            otros_df = reference_df.iloc[0:0].copy()

    # Referencias a hojas, no wb[nombre]: Excel no distingue mayúsculas y openpyxl puede
    # ajustar el título; buscar por string rompe (p. ej. «4-a» vs «4-a1»).
    ws_first = wb[master_name]
    ws_first.title = ordered_sheets[0]
    sheet_list: list = [ws_first]
    for t in ordered_sheets[1:]:
        w = wb.copy_worksheet(sheet_list[0])
        w.title = t
        sheet_list.append(w)

    def fill_sheet_ws(ws, part: pd.DataFrame) -> None:
        label = (config.institution_label or "").strip()
        if label:
            _set_cell_merged_aware(ws, "D2", label)
        ws["B7"] = config.reading_title or ""
        ws["L7"] = config.report_date or ""
        ws["I2"] = config.asignatura or ""
        ws["I3"] = config.docente or ""
        ws["I4"] = config.elaborado_por or ""
        for c_idx, colname in enumerate(header_row, start=1):
            ws.cell(row=8, column=c_idx, value=colname)
        for r_off, (_, row) in enumerate(part.iterrows(), start=9):
            for c_idx, col_name in enumerate(header_row, start=1):
                val = row[col_name]
                if pd.isna(val):
                    val = None
                elif isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
                ws.cell(row=r_off, column=c_idx, value=val)

    for i, t in enumerate(ordered_sheets):
        if t == otros_title:
            fill_sheet_ws(sheet_list[i], otros_df)
        else:
            fill_sheet_ws(sheet_list[i], sheet_frames[t])

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    bio.seek(0)
    return bio.read()


def build_workbook_bytes(
    admin_source: str | Path | BinaryIO,
    template_path: str | Path,
    config: BuildConfig,
    source_filename: str = "",
) -> bytes:
    """
    Carga plantilla + admin y devuelve el .xlsx en memoria (bytes).
    Con BytesIO, pasa `source_filename` (nombre original) para distinguir .csv.
    """
    df, header_row = load_admin_dataframe(admin_source, source_filename=source_filename)
    sheet_frames, otros_df = split_rows_by_sheet(df, config)
    return _compose_excel_bytes(
        Path(template_path), config, header_row, sheet_frames, otros_df, df
    )


def build_workbook_zip_by_institution(
    admin_source: str | Path | BinaryIO,
    template_path: str | Path,
    config: BuildConfig,
    source_filename: str = "",
    institution_filter: list[str] | None = None,
) -> bytes:
    """
    Un .xlsx por valor distinto en «Institución o colegio», empaquetados en un .zip.
    Cada archivo lleva en D2 el nombre de ese colegio.
    `institution_filter`, si se pasa, limita a esos nombres (coincidencia exacta);
    `None` = incluir todos los colegios detectados.
    """
    df, header_row = load_admin_dataframe(admin_source, source_filename=source_filename)
    schools = unique_institution_values(df)
    if not schools:
        raise ValueError(
            "No hay filas con «Institución o colegio» rellenado; no se puede generar un archivo por colegio.",
        )
    if institution_filter is not None:
        if not institution_filter:
            raise ValueError(
                "No hay colegios seleccionados. Marca al menos un centro para el ZIP o no envíes el filtro.",
            )
        wanted = {str(x).strip() for x in institution_filter if str(x).strip() != ""}
        schools = [s for s in schools if s in wanted]
    if not schools:
        raise ValueError(
            "Ningún colegio seleccionado coincide con los del archivo (revisa nombres exactos en «Institución o colegio»).",
        )
    template_path = Path(template_path)
    if not template_path.is_file():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for inst_name in schools:
            sub = df[df[INST_COL].apply(lambda v: str(v).strip() == inst_name)].copy()
            if sub.empty:
                continue
            cfg_s = replace(
                config,
                institution_contains="",
                only_primary_institution=True,
                institution_label=inst_name,
            )
            sheet_frames, otros_df = split_rows_by_sheet(sub, cfg_s)
            xlsx = _compose_excel_bytes(
                template_path, cfg_s, header_row, sheet_frames, otros_df, sub
            )
            zf.writestr(f"informe_lecturas_{_safe_file_stem(inst_name)}.xlsx", xlsx)
    zip_buf.seek(0)
    return zip_buf.read()


def _sheet_sort_key(name: str) -> tuple:
    if name.startswith("G_") and len(name) > 2:
        return _sheet_sort_key(name[2:])
    m = re.match(r"^(\d+)-(.+)$", name)
    if m:
        return (int(m.group(1)), m.group(2))
    return (999, name)


def default_template_path() -> Path:
    base = Path(__file__).resolve().parent.parent
    cand = base / "FORMATO_LEC_AURELIO MARTÍNEZ MUTIS.xlsx"
    return cand
